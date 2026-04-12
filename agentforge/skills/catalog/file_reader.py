"""Skill: file_reader — read local files: text, JSON, CSV, Markdown, PDF, DOCX, XLSX."""
from __future__ import annotations
import json
import csv
import io
from pathlib import Path
from agentforge.skills.base import BaseSkill, SkillInput, SkillOutput

# Supported extensions and their labels
_TEXT_EXTS  = {".txt", ".md", ".py", ".js", ".ts", ".html", ".css",
               ".yaml", ".yml", ".toml", ".env", ".sh", ".rs", ".go", ".java"}
_PARSE_EXTS = {".json", ".csv", ".tsv"}
_BINARY_EXTS = {".pdf", ".docx", ".xlsx", ".xls"}


class FileReaderSkill(BaseSkill):
    name = "file_reader"
    description = (
        "Read local files and return their contents. "
        "Supports plain text/code files, JSON (parsed), CSV/TSV (parsed as rows), "
        "PDF (text extraction), DOCX (paragraphs), and XLSX (sheets as CSV)."
    )
    category = "filesystem"
    tags = ["file", "read", "text", "json", "csv", "pdf", "docx", "xlsx", "markdown"]
    level = "advanced"
    input_schema = {
        "path":         {"type": "string",  "required": True,
                         "description": "Absolute or relative file path"},
        "encoding":     {"type": "string",  "default": "utf-8"},
        "max_bytes":    {"type": "integer", "default": 2_000_000,
                         "description": "Max bytes to read (default 2 MB)"},
        "sheet":        {"type": "string",  "default": "",
                         "description": "XLSX sheet name (empty = first sheet)"},
        "pdf_pages":    {"type": "string",  "default": "all",
                         "description": "PDF page range: 'all' | '1-5' | '3'"},
    }
    output_schema = {
        "content":    {"type": "string",  "description": "Raw text content"},
        "parsed":     {"type": "any",     "description": "Structured data (JSON/CSV/XLSX rows)"},
        "size_bytes": {"type": "integer"},
        "extension":  {"type": "string"},
        "pages":      {"type": "integer", "description": "Page count for PDF"},
        "sheets":     {"type": "array",   "description": "Sheet names for XLSX"},
    }

    async def execute(self, inp: SkillInput) -> SkillOutput:
        file_path = Path(inp.data.get("path", ""))
        encoding  = inp.data.get("encoding", "utf-8")
        max_bytes = int(inp.data.get("max_bytes", 2_000_000))
        sheet     = inp.data.get("sheet", "")
        pdf_pages = inp.data.get("pdf_pages", "all")

        if not file_path.exists():
            return SkillOutput.fail(f"File not found: {file_path}")
        if not file_path.is_file():
            return SkillOutput.fail(f"Path is not a file: {file_path}")

        ext = file_path.suffix.lower()
        size_bytes = file_path.stat().st_size
        content = ""
        parsed  = None
        pages   = 0
        sheets: list[str] = []

        # ── PDF ──────────────────────────────────────────────────────────
        if ext == ".pdf":
            try:
                import pdfplumber
                with pdfplumber.open(file_path) as pdf:
                    pages = len(pdf.pages)
                    page_range = self._parse_page_range(pdf_pages, pages)
                    text_parts = []
                    for i in page_range:
                        p = pdf.pages[i]
                        t = p.extract_text() or ""
                        text_parts.append(f"--- Page {i+1} ---\n{t}")
                content = "\n\n".join(text_parts)[:max_bytes]
            except ImportError:
                return SkillOutput.fail("pdfplumber not installed. Run: pip install pdfplumber")

        # ── DOCX ──────────────────────────────────────────────────────────
        elif ext == ".docx":
            try:
                from docx import Document
                doc = Document(file_path)
                content = "\n".join(p.text for p in doc.paragraphs if p.text.strip())[:max_bytes]
            except ImportError:
                return SkillOutput.fail("python-docx not installed. Run: pip install python-docx")

        # ── XLSX / XLS ────────────────────────────────────────────────────
        elif ext in (".xlsx", ".xls"):
            try:
                import openpyxl
                wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
                sheets = wb.sheetnames
                ws = wb[sheet] if sheet and sheet in wb.sheetnames else wb.active
                rows = []
                for row in ws.iter_rows(values_only=True):
                    rows.append([str(c) if c is not None else "" for c in row])
                # Convert to CSV string
                buf = io.StringIO()
                writer = csv.writer(buf)
                writer.writerows(rows)
                content = buf.getvalue()[:max_bytes]
                parsed  = rows[:200]  # First 200 rows as list-of-lists
            except ImportError:
                return SkillOutput.fail("openpyxl not installed. Run: pip install openpyxl")

        # ── JSON ──────────────────────────────────────────────────────────
        elif ext == ".json":
            raw = file_path.read_bytes()[:max_bytes]
            content = raw.decode(encoding, errors="replace")
            try:
                parsed = json.loads(content)
            except json.JSONDecodeError as e:
                return SkillOutput.fail(f"Invalid JSON: {e}")

        # ── CSV / TSV ─────────────────────────────────────────────────────
        elif ext in (".csv", ".tsv"):
            raw = file_path.read_bytes()[:max_bytes]
            content = raw.decode(encoding, errors="replace")
            delim = "\t" if ext == ".tsv" else ","
            try:
                reader = csv.DictReader(io.StringIO(content), delimiter=delim)
                parsed = list(reader)
            except Exception:
                pass

        # ── Plain text / code ─────────────────────────────────────────────
        else:
            raw     = file_path.read_bytes()[:max_bytes]
            content = raw.decode(encoding, errors="replace")

        return SkillOutput(
            data={
                "content":    content,
                "parsed":     parsed,
                "size_bytes": size_bytes,
                "extension":  ext,
                "pages":      pages,
                "sheets":     sheets,
            }
        )

    @staticmethod
    def _parse_page_range(spec: str, total: int) -> range:
        spec = spec.strip().lower()
        if spec == "all":
            return range(total)
        if "-" in spec:
            parts = spec.split("-", 1)
            start = int(parts[0]) - 1
            end   = int(parts[1])
            return range(max(0, start), min(end, total))
        return range(int(spec) - 1, int(spec))  # single page
